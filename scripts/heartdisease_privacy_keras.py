""" IMPORTING LIBRARIES AND LOADING DATA """
import datetime
import os
import sys
import dp_accounting
import openpyxl
import pandas as pd
import numpy as np
from IPython.display import display
import matplotlib.pyplot as plt
# %matplotlib inline
from absl import flags
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
import tensorflow as tf
from tensorflow import estimator as tf_estimator
from tensorflow_privacy import DPKerasSGDOptimizer
from tensorflow_privacy.privacy.keras_models import dp_keras_model
import wandb
from environs import Env
from tensorflow_privacy.privacy.optimizers.dp_optimizer_keras_sparse import DPSparseKerasSGDOptimizer

from scripts.custom_utils import log_metrics_dp, log_metrics_baseline, compute_epsilon_noise

""" Environmental variables """
env = Env()

""" Directories and timestamps """
actual_dir = None
excel_path = None
timestamp = None
logger = None

excel_file_name = "metrics.xlsx"
metrics_columns = ['model', 'optimizer', 'epsilon', 'noise_multiplier', 'l2_norm_clip', 'batch_size', 'microbatches',
                   'learning_rate', 'acc', 'val_acc', 'loss', 'val_loss']
""" Training parameters """
total_steps = 0
steps_per_epoch = 0
total_samples = 0
""" Hyperparameters """

flags.DEFINE_float('learning_rate', 0.0001, 'Learning rate for training')
flags.DEFINE_float('noise_multiplier', 1,
                   'Ratio of the standard deviation to the clipping norm')
flags.DEFINE_float('l2_norm_clip', 0.6, 'Clipping norm')
flags.DEFINE_integer('batch_size', 25, 'Batch size')
flags.DEFINE_integer('epochs', 6, 'Number of epochs')
flags.DEFINE_integer(
    'microbatches', 25, 'Number of microbatches '
                        '(must evenly divide batch_size)')

FLAGS = flags.FLAGS
FLAGS(sys.argv)

""" Create and compile model"""


def get_layers_Binary_Classification():
    return [tf.keras.layers.InputLayer(input_shape=(39,)),
            tf.keras.layers.Dropout(0.2),
            tf.keras.layers.Dense(5, activation='relu'),
            tf.keras.layers.Dropout(0.2),
            tf.keras.layers.Dense(1, activation='sigmoid')]


def get_layers_Linear_Regression():
    return [tf.keras.layers.Dense(1, activation="linear")]


def create_baseline_models():
    model = []

    """Regular Binary Classification Baseline"""
    model_baseline_binary = tf.keras.Sequential(
        get_layers_Binary_Classification())

    optimizer = tf.keras.optimizers.SGD(learning_rate=FLAGS.learning_rate)

    model_baseline_binary.compile(optimizer=optimizer,
                                  loss='mse',
                                  metrics='accuracy')

    model_baseline_linear = tf.keras.Sequential(
        get_layers_Linear_Regression())

    optimizer = tf.keras.optimizers.Adam(learning_rate=FLAGS.learning_rate)

    model_baseline_linear.compile(optimizer=optimizer,
                                  loss='mean_squared_error',
                                  metrics='accuracy')

    model.append(model_baseline_binary)
    model.append(model_baseline_linear)
    return model


def get_preprocessed_data():
    df = pd.read_csv(os.environ['DATASET_PATH'])

    logger.log({"dataset": wandb.Table(dataframe=df)})

    print(df.LastCheckupTime.value_counts())
    print(df.State.value_counts())
    print(df.HadHeartAttack.unique)

    y = df.HadHeartAttack.replace(['Yes', 'No'], [1, 0])

    x = df.drop('HadHeartAttack', axis=1)
    pd.set_option('display.max_columns', None)
    display(x)

    # encode categorical data to numerical
    enc = LabelEncoder()
    for i in x.columns:
        if x[i].dtype == 'object':
            x[i] = enc.fit_transform(x[i])
    print(x.info())

    x.drop(axis=0, index=x.index[-22:], inplace=True)
    y.drop(axis=0, index=y.index[-22:], inplace=True)

    return x, y


""" Grid search approach"""


def calculate_model(x_train, x_test, y_train, y_test, m, index):
    if FLAGS.batch_size % FLAGS.microbatches != 0:
        raise ValueError('Batch size should be an integer multiple of the number of microbatches')

    """ Binary classification"""
    print("Training with batch_size = " + str(FLAGS.batch_size) +
          ", microbatches = " + str(FLAGS.microbatches) +
          ", l2_norm_clip = " + str(m["l2_norm_clip"]) +
          ", noise_multiplier = " + str(m["noise_multiplier"]) +
          ", learning_rate = " + str(FLAGS.learning_rate))
    model = tf.keras.Sequential()
    optimizer = tf.keras.optimizers.SGD(learning_rate=FLAGS.learning_rate)
    loss = tf.keras.losses.BinaryCrossentropy(from_logits=False)
    history = tf.keras.callbacks.History()

    if m["optimizer"] == '0':
        model = tf.keras.Sequential(
            [
                tf.keras.layers.Dense(5, activation='relu'),
                tf.keras.layers.Dropout(0.2),
                tf.keras.layers.Dense(1, activation='sigmoid')])
        optimizer = DPKerasSGDOptimizer(
            l2_norm_clip=m["l2_norm_clip"],
            noise_multiplier=m["noise_multiplier"],
            num_microbatches=FLAGS.microbatches,
            learning_rate=FLAGS.learning_rate
        )
        loss = tf.keras.losses.MeanSquaredError(reduction='none')
        m["optimizer"] = "SGD-DP"
        m['model'] = 'SGD'
    elif m["optimizer"] == '1':
        model = tf.keras.Sequential(
            [
                tf.keras.layers.Dense(5, activation='relu'),
                tf.keras.layers.Dropout(0.2),
                tf.keras.layers.Dense(1, activation='sigmoid')])
        optimizer = DPSparseKerasSGDOptimizer(
            l2_norm_clip=m["l2_norm_clip"],
            noise_multiplier=m["noise_multiplier"],
            num_microbatches=FLAGS.microbatches,
            learning_rate=FLAGS.learning_rate
        )
        m["optimizer"] = "SGD-SPARSE-DP"
        m['model'] = 'SGD'
        loss = tf.keras.losses.MeanAbsoluteError(
            reduction=tf.keras.losses.Reduction.NONE
        )
    elif m["optimizer"] == '2':
        model = dp_keras_model.DPSequential(
            l2_norm_clip=m["l2_norm_clip"],
            noise_multiplier=m["noise_multiplier"],
            num_microbatches=FLAGS.microbatches,
            layers=get_layers_Binary_Classification())
        optimizer = tf.keras.optimizers.SGD(learning_rate=FLAGS.learning_rate)
        m["optimizer"] = "SGD-NO-DP"
        m['model'] = 'SGD-DP'

    model.compile(optimizer=optimizer, loss=loss, metrics=['accuracy'])
    history = model.fit(x_train, y_train,
                        epochs=FLAGS.epochs,
                        validation_data=(x_test, y_test),
                        batch_size=FLAGS.batch_size)
    m['acc'] = round(history.history['accuracy'][-1], 4)
    m['val_acc'] = round(history.history['val_accuracy'][-1], 4)
    m['loss'] = round(history.history['loss'][-1], 4)
    m['val_loss'] = round(history.history['val_loss'][-1], 4)

    print("Best metrics are: acc = " + str(history.history['accuracy'][-1]) +
          ", val_acc = " + str(history.history['val_accuracy'][-1]) +
          ", loss = " + str(history.history['loss'][-1]) +
          ", val_loss = " + str(history.history['val_loss'][-1]))

    log_metrics_dp(history, m, index, actual_dir, m['optimizer'])


def main():
    global timestamp, actual_dir, logger, excel_path, total_steps, steps_per_epoch, total_samples
    """ Initialize wandb"""
    logger = wandb.init(
        # set the wandb project where this run will be logged
        project="Hyperparameter Tuning in Privacy-Preserving Machine Learning",

        # track hyperparameters and run metadata
        config={
            "learning_rate": FLAGS.learning_rate,
            "dataset": "heart_2022_no_nans",
            "epochs": FLAGS.epochs,
        }
    )

    """ Create directories for plots """
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + "/"
    actual_dir = os.path.join("../Plots/", timestamp)
    os.makedirs(actual_dir)

    """ Create directory for metrics"""
    excel_dir = os.path.join("../", "Metrics/")
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    # setup excel_path
    excel_path = os.path.join(excel_dir, excel_file_name)
    book = openpyxl.load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book

    """ Prepare and split data"""
    x, y = get_preprocessed_data()
    # Take a look at the number of rows
    print("Data shape: " + str(x.shape))
    print("Labels shape: " + str(y.shape))

    x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2682926)
    print("X_train shape: " + str(x_train.shape))
    print("X_test shape: " + str(x_test.shape))

    total_samples = x_train.shape[0]
    total_steps = int(np.ceil(total_samples / FLAGS.batch_size)) * FLAGS.epochs
    steps_per_epoch = int(np.ceil(total_steps / FLAGS.epochs))

    """Get baseline results"""
    baseline_models = create_baseline_models()
    baseline_results = pd.DataFrame(columns=metrics_columns)
    for index, model in enumerate(baseline_models):
        history = model.fit(x_train, y_train,
                            epochs=FLAGS.epochs,
                            validation_data=(x_test, y_test),
                            batch_size=FLAGS.batch_size)
        if index == 0:
            m = "DNN-NO-DP"
        else:
            m = "LR-NO-DP"
        new_row = {'model': m,
                   'optimizer': 'SGD-NO-DP',
                   'epsilon': "INFINITY",
                   'noise_multiplier': "0",
                   'l2_norm_clip': "0",
                   'batch_size': FLAGS.batch_size,
                   'microbatches': FLAGS.microbatches,
                   'learning_rate': FLAGS.learning_rate,
                   'acc': round(history.history['accuracy'][-1], 4),
                   'val_acc': round(history.history['val_accuracy'][-1], 4),
                   'loss': round(history.history['loss'][-1], 4),
                   'val_loss': round(history.history['val_loss'][-1], 4),
                   }

        log_metrics_baseline(history, index, actual_dir, FLAGS.batch_size, FLAGS.microbatches)
        baseline_results.loc[len(baseline_results)] = new_row
    if "Baseline" not in book.sheetnames:
        book.create_sheet("Baseline")
    baseline_results.to_excel(writer, sheet_name="Baseline", startrow=writer.sheets["Baseline"].max_row,
                              index=True, header=False)

    """ Grid Search """
    model_grid = []

    aVals = [.25, .5, 1, 1.25, 1.5, 2]

    for optimizer_index in range(0, 3):
        for norm_clip in aVals:
            for noise_mul in aVals:
                model_grid.append({
                    'l2_norm_clip': FLAGS.l2_norm_clip * norm_clip,
                    'noise_multiplier': FLAGS.noise_multiplier * noise_mul,
                    'epochs': FLAGS.epochs,
                    'acc': 0,
                    'val_acc': 0,
                    'loss': 0,
                    'val_loss': 0,
                    'optimizer': str(optimizer_index),
                    'model': 'model'
                })

    for i, m in enumerate(model_grid):
        calculate_model(
            x_train, x_test, y_train, y_test, m, i
        )

    # calculate epsilon
    if env.bool("DP", False) is True:
        privacy_results = pd.DataFrame(columns=metrics_columns)
        for m in model_grid:
            epsilon = compute_epsilon_noise(FLAGS.epochs * x_train.shape[0] // FLAGS.batch_size, m['noise_multiplier'],
                                            FLAGS.batch_size, total_samples)
            new_row = {'model': m['model'],
                       'optimizer': m['optimizer'],
                       'epsilon': epsilon,
                       'noise_multiplier': m['noise_multiplier'],
                       'l2_norm_clip': m['l2_norm_clip'],
                       'batch_size': FLAGS.batch_size,
                       'microbatches': FLAGS.microbatches,
                       'learning_rate': FLAGS.learning_rate,
                       'acc': m['acc'],
                       'val_acc': m['val_acc'],
                       'loss': m['loss'],
                       'val_loss': m['val_loss'],
                       }

            privacy_results.loc[len(privacy_results)] = new_row
            print("Computed epsilon with l2_norm_clip = " + str(m['l2_norm_clip']) + ", noise_multiplier = " +
                  str(m['noise_multiplier']) + " is epsilon = " + str(epsilon))

        models = privacy_results['model'].unique()
        optimizers = privacy_results['optimizer'].unique()
        for m in models:
            for o in optimizers:
                if m + "_" + o not in book.sheetnames:
                    book.create_sheet(m + "_" + o)
                actual = privacy_results[(privacy_results["model"] == m) & (privacy_results["optimizer"] == o)]
                actual.to_excel(writer, sheet_name=m, startrow=writer.sheets[m].max_row, index=True, header=True)

        writer.save()

        l2_norm_clip_values = privacy_results['l2_norm_clip'].unique()
        print(l2_norm_clip_values)
        for u in l2_norm_clip_values:
            actual = privacy_results[privacy_results["l2_norm_clip"] == u]
            # sorted_actual = sorted(actual, key=lambda x: x["val_acc"], reverse=True)
            print(actual)
            # plot epsilon graph
            plt.plot(actual["noise_multiplier"], actual["epsilon"])
            plt.title('model epsilon values for l2_norm_clip ' + str(u))
            plt.ylabel('epsilon')
            plt.xlabel('noise_multiplier')
            plt.grid(True)
            plt.tight_layout()
            plt.savefig(actual_dir + "epsilon_l2_norm_clip_" + str(u).replace(".", '_'), bbox_inches='tight')
            logger.log({"Epsilon plot l2 = " + str(u): plt})
            plt.show()
            # plot accuracy graph
            plt.plot(actual["noise_multiplier"], actual["val_acc"])
            plt.plot(actual["noise_multiplier"], actual["val_loss"])
            plt.title('model val_acc, val_loss values for l2_norm_clip ' + str(u))
            plt.ylabel('val_acc, val_loss')
            plt.xlabel('noise_multiplier')
            plt.grid(True)
            plt.tight_layout()
            plt.savefig(actual_dir + "noise_accuracy_l2_norm_clip_" + str(u).replace(".", '_'), bbox_inches='tight')
            logger.log({"Accuracy by noise and clip plot l2 = " + str(u): plt})
            plt.show()

    wandb.finish()


if __name__ == '__main__':
    main()
