import os

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt


def plot_acc(df_noise, excel_sheet_dir):
    plt.plot(df_noise['noise_multiplier'], df_noise['avg_acc'])
    for noise in df_noise['noise_multiplier'].unique():
        avg_acc = df_noise[df_noise['noise_multiplier'] == noise].avg_acc.tolist()[0]
        plt.text(noise,
                 avg_acc,
                 f'{round(avg_acc, 4)}', ha='right')

    plt.plot(df_noise['noise_multiplier'], df_noise['avg_val_acc'])
    for noise in df_noise['noise_multiplier'].unique():
        avg_val_acc = df_noise[df_noise['noise_multiplier'] == noise].avg_val_acc.tolist()[0]
        plt.text(noise,
                 avg_val_acc,
                 f'{round(avg_val_acc, 4)}', ha='right')

    plt.title(
        df_noise['model'][0] + ' model (' + df_noise['optimizer'][0] + ') averages with l2_norm_clip: ' +
        str(df_noise['l2_norm_clip'][0]))
    plt.ylabel('acc')
    plt.xlabel('noise_multiplier')
    plt.legend(['avg_acc', 'avg_val_acc'], loc='upper left')
    plt.grid(True)

    # save png of the plot
    plt.savefig(str(excel_sheet_dir) + "l2_" + str(df_noise['l2_norm_clip'][0]).replace(".", "_") + "_acc",
                bbox_inches='tight')
    plt.show()


def plot_time(df_noise, excel_sheet_dir):
    plt.plot(df_noise['noise_multiplier'], df_noise['avg_training_time'])
    for noise in df_noise['noise_multiplier'].unique():
        time = df_noise[df_noise['noise_multiplier'] == noise].avg_training_time.tolist()[0]
        plt.text(noise,
                 time,
                 f'{round(time, 4)}', ha='right')

    plt.title(
        df_noise['model'][0] + ' model (' + df_noise['optimizer'][0] + ') averages with l2_norm_clip: ' +
        str(df_noise['l2_norm_clip'][0]))
    plt.ylabel('seconds')
    plt.xlabel('noise_multiplier')
    plt.legend(['avg_training_time'], loc='upper left')
    plt.grid(True)

    # save png of the plot
    plt.savefig(str(excel_sheet_dir) + "l2_" + str(df_noise['l2_norm_clip'][0]).replace(".", "_") + "_time",
                bbox_inches='tight')
    plt.show()


def plot_loss(df_noise, excel_sheet_dir):
    plt.plot(df_noise['noise_multiplier'], df_noise['avg_loss'])
    for noise in df_noise['noise_multiplier'].unique():
        avg_loss = df_noise[df_noise['noise_multiplier'] == noise].avg_loss.tolist()[0]
        plt.text(noise,
                 avg_loss,
                 f'{round(avg_loss, 4)}', ha='right')

    plt.plot(df_noise['noise_multiplier'], df_noise['avg_val_loss'])
    for noise in df_noise['noise_multiplier'].unique():
        avg_val_loss = df_noise[df_noise['noise_multiplier'] == noise].avg_val_loss.tolist()[0]
        plt.text(noise,
                 avg_val_loss,
                 f'{round(avg_val_loss, 4)}', ha='right')

    plt.title(
        df_noise['model'][0] + ' model (' + df_noise['optimizer'][0] + ') averages with l2_norm_clip: ' +
        str(df_noise['l2_norm_clip'][0]))
    plt.ylabel('loss')
    plt.xlabel('noise_multiplier')
    plt.legend(['avg_loss', 'avg_val_loss'], loc='upper left')
    plt.grid(True)

    # save png of the plot
    plt.savefig(str(excel_sheet_dir) + "l2_" + str(df_noise['l2_norm_clip'][0]).replace(".", "_") + "_loss",
                bbox_inches='tight')
    plt.show()


def main():
    metrics_avgs_dir = '../Metrics/Metrics_Averages/'
    if not os.path.exists(metrics_avgs_dir):
        os.makedirs(metrics_avgs_dir)

    # Specify the path to your Excel file
    excel_in_file_path = '../Metrics/metrics.xlsx'
    excel_dir = os.path.join("../", "Metrics/")
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    # setup excel_path
    excel_path = os.path.join(metrics_avgs_dir, "metrics_avgs.xlsx")
    book = openpyxl.load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    writer.book = book

    # Get the list of all sheet names in the Excel file
    all_sheet_names = pd.ExcelFile(excel_in_file_path).sheet_names

    # Specify the columns you want to read
    res_columns = ['model', 'optimizer', 'epsilon', 'noise_multiplier', 'l2_norm_clip', 'batch_size',
                   'microbatches', 'learning_rate',
                   'avg_training_time', 'avg_acc', 'avg_val_acc', 'avg_loss', 'avg_val_loss']



    for sheet_name in all_sheet_names:
        df_res = pd.DataFrame(columns=res_columns)
        if sheet_name == 'Baseline':
            continue
        df = pd.read_excel(excel_in_file_path, sheet_name=sheet_name)
        unique_noises = df['noise_multiplier'].unique()
        unique_clips = df['l2_norm_clip'].unique()
        excel_sheet_dir = os.path.join(excel_dir, sheet_name + "Plots/")
        if not os.path.exists(excel_sheet_dir):
            os.makedirs(excel_sheet_dir)
        for unique_clip in unique_clips:
            df_noise = pd.DataFrame(columns=res_columns)
            for unique_noise in unique_noises:
                actual_rows = df[(df['noise_multiplier'] == unique_noise) &
                                 (df['l2_norm_clip'] == unique_clip)]
                if len(actual_rows) != 0:
                    num_rows = len(actual_rows)

                    new_row = {'model': actual_rows['model'].tolist()[0],
                               'optimizer': actual_rows['optimizer'].tolist()[0],
                               'epsilon': actual_rows['epsilon'].tolist()[0],
                               'noise_multiplier': actual_rows['noise_multiplier'].tolist()[0],
                               'l2_norm_clip': round(actual_rows['l2_norm_clip'].tolist()[0], 3),
                               'batch_size': actual_rows['batch_size'].tolist()[0],
                               'microbatches': actual_rows['microbatches'].tolist()[0],
                               'learning_rate': actual_rows['learning_rate'].tolist()[0],
                               'avg_training_time': sum(actual_rows['training_time']) / num_rows,
                               'avg_acc': sum(actual_rows['acc']) / num_rows,
                               'avg_val_acc': sum(actual_rows['val_acc']) / num_rows,
                               'avg_loss': sum(actual_rows['loss']) / num_rows,
                               'avg_val_loss': sum(actual_rows['val_loss']) / num_rows,
                               }

                    df_noise.loc[len(df_noise)] = new_row
                    df_res.loc[len(df_res)] = new_row

            # Plot the averages for hyperparams

            plot_acc(df_noise, excel_sheet_dir)
            plot_loss(df_noise, excel_sheet_dir)
            plot_time(df_noise, excel_sheet_dir)

        if sheet_name not in book.sheetnames:
            book.create_sheet(sheet_name)
        df_res.to_excel(writer, sheet_name=sheet_name, startrow=writer.sheets[sheet_name].max_row, index=True,
                        header=True)
    writer.save()
    return


if __name__ == '__main__':
    main()
