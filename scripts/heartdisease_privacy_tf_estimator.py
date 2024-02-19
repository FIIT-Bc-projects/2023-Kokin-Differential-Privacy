""" IMPORTING LIBRARIES AND LOADING DATA """
import datetime
import os
import sys
import dp_accounting
import openpyxl
import pandas as pd
from IPython.display import display
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
# %matplotlib inline
import seaborn as sns
import missingno as msno
import tensorflow_privacy
from absl import flags
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
import tensorflow as tf
from tensorflow import estimator as tf_estimator
from tensorflow_privacy import DPKerasSGDOptimizer
from tensorflow_privacy.privacy.optimizers import dp_optimizer
import wandb
from environs import Env
from tensorflow_privacy.privacy.optimizers.dp_optimizer import DPAdamGaussianOptimizer, DPGradientDescentOptimizer
from batch_utils import MetricHistoryCallback
from scripts.dataset_utils import get_preprocessed_data
from scripts.dp_utils import compute_epsilon_noise, create_baseline_models

""" Environmental variables """
env = Env()

""" Directories and timestamps """
actual_dir = None
excel_path = None
timestamp = None
logger = None

excel_file_name = "metrics.xlsx"
metrics_columns = ['model', 'optimizer', 'epsilon', 'noise_multiplier', 'l2_norm_clip', 'acc', 'val_acc', 'loss',
                   'val_loss', 'batch_size', 'microbatches', 'learning_rate']

""" Hyperparameters """

flags.DEFINE_float('learning_rate', 0.0001, 'Learning rate for training')
flags.DEFINE_float('noise_multiplier', 1,
                   'Ratio of the standard deviation to the clipping norm')
flags.DEFINE_float('l2_norm_clip', 1, 'Clipping norm')
flags.DEFINE_integer('batch_size', 50, 'Batch size')
flags.DEFINE_integer('epochs', 6, 'Number of epochs')
flags.DEFINE_integer(
    'microbatches', 50, 'Number of microbatches '
                        '(must evenly divide batch_size)')

FLAGS = flags.FLAGS
FLAGS(sys.argv)


# Define an input function for training
def train_input_fn(features, labels, shuffle=True):
    dataset = tf.data.Dataset.from_tensor_slices(({"feature": features}, labels))
    if shuffle:
        dataset = dataset.shuffle(buffer_size=10000)
    dataset = dataset.batch(FLAGS.batch_size).repeat(FLAGS.epochs)
    return dataset


# Define an input function for validation
def val_input_fn(features, labels, batch_size):
    dataset = tf.data.Dataset.from_tensor_slices(({"feature": features}, labels))
    dataset = dataset.batch(batch_size)
    return dataset


def dnn_dp_model(features, labels, mode, params):
    """Model function for a DNN."""
    # Define DNN architecture using tf.keras.layers.
    input_layer = tf.keras.layers.InputLayer(input_shape=(features.shape[1],))

    hidden1 = tf.keras.layers.Dense(5, activation='relu')(input_layer)
    dropout1 = tf.keras.layers.Dropout(0.2)(hidden1)

    hidden2 = tf.keras.layers.Dense(3, activation='relu')(dropout1)
    dropout2 = tf.keras.layers.Dropout(0.2)(hidden2)

    logits = tf.keras.layers.Dense(1, activation='sigmoid')(dropout2)

    predictions = tf.nn.sigmoid(logits)
    predicted_classes = tf.round(predictions)

    # Calculate loss as a vector (to support microbatches in DP-SGD).
    vector_loss = tf.nn.sigmoid_cross_entropy_with_logits(
        labels=labels, logits=logits)
    # Define mean of loss across minibatch (for reporting through tf.Estimator).
    scalar_loss = tf.reduce_mean(input_tensor=vector_loss)

    if mode == tf.estimator.ModeKeys.TRAIN:
        optimizer = "adam"
        train_op = None
        callback = MetricHistoryCallback()
        if params['optimizer'] == '0':
            optimizer = tf.compat.v1.train.GradientDescentOptimizer(learning_rate=FLAGS.learning_rate)
            params['optimizer'] = "SGD-NO-DP"
            train_op = optimizer.minimize(scalar_loss, global_step=tf.compat.v1.train.get_global_step())
        elif params['optimizer'] == '1':
            optimizer = DPGradientDescentOptimizer(
                l2_norm_clip=params["l2_norm_clip"],
                noise_multiplier=params["noise_multiplier"],
                num_microbatches=FLAGS.microbatches,
                learning_rate=FLAGS.learning_rate)
            params['optimizer'] = "SGD-DP"
            train_op = optimizer.minimize(scalar_loss)
        elif params['optimizer'] == '2':
            optimizer = DPAdamGaussianOptimizer(
                l2_norm_clip=params["l2_norm_clip"],
                noise_multiplier=params["noise_multiplier"],
                num_microbatches=FLAGS.microbatches,
                learning_rate=FLAGS.learning_rate)
            params['optimizer'] = "ADAM-DP"
            # Wrap the optimizer with privacy computations
            train_op = optimizer.minimize(scalar_loss, global_step=tf.compat.v1.train.get_global_step())

        return tf.estimator.EstimatorSpec(
            mode, loss=scalar_loss, train_op=train_op, training_hooks=[callback]
        )

    if mode == tf.estimator.ModeKeys.EVAL:
        accuracy = tf.compat.v1.metrics.accuracy(labels=labels, predictions=predicted_classes)
        val_loss = tf.reduce_sum(tf.nn.sigmoid_cross_entropy_with_logits(labels=labels, logits=logits)) / tf.cast(
            tf.shape(labels)[0], tf.float32)
        val_metric_ops = {'accuracy': accuracy, 'val_loss': val_loss}
        return tf.estimator.EstimatorSpec(mode, loss=scalar_loss, eval_metric_ops=val_metric_ops)



""" Grid search approach"""


def calculate_model(x_train, x_test, y_train, y_test, m, index):
    if FLAGS.batch_size % FLAGS.microbatches != 0:
        raise ValueError('Batch size should be an integer multiple of the number of microbatches')

    """ Binary classification"""
    estimator = tf.estimator.Estimator(
        model_fn=dnn_dp_model,
        params=m,
    )

    # Extract column names from the DataFrame
    feature_columns = x_train.columns.tolist()
    # Create numeric feature columns for each column in the DataFrame
    numeric_feature_columns = [tf.feature_column.numeric_column(key=feature) for feature in feature_columns]
    input_dim = x_train.shape[1]
    numeric_feature_columns = [tf.feature_column.numeric_column(key=feature, shape=(input_dim,)) for feature in
                               feature_columns]

    print("Training with batch_size = " + str(FLAGS.batch_size) +
          ", microbatches = " + str(FLAGS.microbatches) +
          ", l2_norm_clip = " + str(m["l2_norm_clip"]) +
          ", noise_multiplier = " + str(m["noise_multiplier"]) +
          ", learning_rate = " + str(FLAGS.learning_rate))

    for epoch in range(FLAGS.epochs):
        # Train the Estimator
        estimator.train(
            input_fn=lambda: train_input_fn(x_train, y_train, shuffle=True)
        )

        # Evaluate the Estimator on the validation set
        eval_result = estimator.evaluate(
            input_fn=lambda: val_input_fn(x_test, y_test, batch_size=FLAGS.batch_size)
        )

        print(f"Epoch {epoch + 1}/{FLAGS.epochs}")
        print("Validation Loss:", eval_result['loss'])
        print("Validation Accuracy:", eval_result['accuracy'])

    m['acc'] = round(history.history['accuracy'][-1], 4)
    m['val_acc'] = round(history.history['val_accuracy'][-1], 4)
    m['loss'] = round(history.history['loss'][-1], 4)
    m['val_loss'] = round(history.history['val_loss'][-1], 4)

    print("Best metrics are: acc = " + str(history.history['accuracy'][-1]) +
          ", val_acc = " + str(history.history['val_accuracy'][-1]) +
          ", loss = " + str(history.history['loss'][-1]) +
          ", val_loss = " + str(history.history['val_loss'][-1]))

    log_metrics_dp(history, m, index)


def log_metrics_dp(history, m, index):
    history_dict = history.history
    history_df = pd.DataFrame(history_dict)

    # summarize history for metrics
    plt.plot(history.history['accuracy'])
    for i, metric in enumerate(history.history['accuracy']):
        plt.text(i, history.history['accuracy'][i], f'{round(metric, 4)}', ha='right')

    plt.plot(history.history['val_accuracy'])
    for i, metric in enumerate(history.history['val_accuracy']):
        plt.text(i, history.history['val_accuracy'][i], f'{round(metric, 4)}', ha='left')

    plt.plot(history.history['loss'])
    for i, metric in enumerate(history.history['loss']):
        plt.text(i, history.history['loss'][i], f'{round(metric, 4)}', ha='left')

    plt.plot(history.history['val_loss'])
    for i, metric in enumerate(history.history['val_loss']):
        plt.text(i, history.history['val_loss'][i], f'{round(metric, 4)}', ha='right')

    plt.title(
        'DP Model acc and loss with noise: ' + str(m['noise_multiplier']) + ', l2_norm_clip: ' + str(m['l2_norm_clip']))
    plt.ylabel('loss, acc')
    plt.xlabel('epoch')
    plt.legend(['train_acc', 'test_acc', 'train_loss', 'test_loss'], loc='upper left')
    plt.grid(True)
    plt.tight_layout()

    # save png of the plot
    plt.savefig(str(actual_dir) + "dp_n_" + str(m['noise_multiplier']).replace(".", "_")
                + "l2_" + str(m['l2_norm_clip']).replace(".", "_"), bbox_inches='tight')
    plt.show()

    title = "Accuracy and Loss metrics of DP model training: noise=%s, l2_norm_clip=%s" % (str(m['noise_multiplier']),
                                                                                           str(m['l2_norm_clip']))

    wandb.log(
        {"Accuracy and Loss metrics of DP model (" + str(index) + ")":
            wandb.plot.line_series(
                xs=history_df.index,
                ys=[history_df["loss"],
                    history_df["val_loss"],
                    history_df["accuracy"],
                    history_df["val_accuracy"]
                    ],
                keys=["loss", "val_loss", "accuracy", "val_accuracy"],
                title=title,
                xname="epochs")})


def log_metrics_baseline(history, index):
    history_dict = history.history
    history_df = pd.DataFrame(history_dict)
    # summarize history for accuracy
    plt.plot(history.history['accuracy'])
    for i, metric in enumerate(history.history['accuracy']):
        plt.text(i, history.history['accuracy'][i], f'{round(metric, 4)}', ha='right')

    plt.plot(history.history['val_accuracy'])
    for i, metric in enumerate(history.history['val_accuracy']):
        plt.text(i, history.history['val_accuracy'][i], f'{round(metric, 4)}', ha='left')

    plt.plot(history.history['loss'])
    for i, metric in enumerate(history.history['loss']):
        plt.text(i, history.history['loss'][i], f'{round(metric, 4)}', ha='left')

    plt.plot(history.history['val_loss'])
    for i, metric in enumerate(history.history['val_loss']):
        plt.text(i, history.history['val_loss'][i], f'{round(metric, 4)}', ha='right')

    plt.ylabel('loss, acc')
    plt.xlabel('epoch')
    plt.legend(['train', 'test'], loc='upper left')
    plt.grid(True)
    plt.tight_layout()
    if index == 0:
        plt.title('Binary model Accuracy and Loss metrics with batch_size = ' + str(FLAGS.batch_size) +
                  ', microbatches = ' + str(FLAGS.microbatches) + ", ")

        plt.savefig(str(actual_dir) + "baseline_binary_acc", bbox_inches='tight')
        wandb.log(
            {"Accuracy and Loss metrics of baseline Binary Classification model":
                wandb.plot.line_series(
                    xs=history_df.index,
                    ys=[history_df["accuracy"],
                        history_df["val_accuracy"]],
                    keys=["accuracy", "val_accuracy"],
                    title="Accuracy metrics of baseline Binary Classification model",
                    xname="epochs"
                )})

    else:
        plt.title('Linear model Accuracy and Loss metrics with batch_size = ' + str(FLAGS.batch_size) +
                  ', microbatches = ' + str(FLAGS.microbatches) + ", ")

        plt.savefig(str(actual_dir) + "baseline_linear_acc", bbox_inches='tight')
        wandb.log(
            {"Accuracy and Loss metrics of baseline Binary Classification model":
                wandb.plot.line_series(
                    xs=history_df.index,
                    ys=[history_df["accuracy"],
                        history_df["val_accuracy"]],
                    keys=["accuracy", "val_accuracy"],
                    title="Accuracy metrics of baseline Linear regression model",
                    xname="epochs")})
    plt.show()


def main():
    global timestamp, actual_dir, logger, excel_path
    """ Initialize wandb"""
    logger = wandb.init(
        # set the wandb project where this run will be logged
        project="Hyperparameter Tuning in Privacy-Preserving Machine Learning",

        # track hyperparameters and run metadata
        config={
            "learning_rate": FLAGS.learning_rate,
            "dataset": "heart_2022_no_nans",
            "epochs": FLAGS.epochs,
        }
    )

    """ Create directories for plots """
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + "/"
    actual_dir = os.path.join("../Plots/", timestamp)
    os.makedirs(actual_dir)

    """ Create directory for metrics"""
    excel_dir = os.path.join("../", "Metrics/")
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    # setup excel_path
    excel_path = os.path.join(excel_dir, excel_file_name)
    book = openpyxl.load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book

    """ Prepare and split data"""
    x, y = get_preprocessed_data(logger, wandb)
    # Take a look at the number of rows
    print("Data shape: " + str(x.shape))
    print("Labels shape: " + str(y.shape))

    x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2682926)
    print("X_train shape: " + str(x_train.shape))
    print("X_test shape: " + str(x_test.shape))

    """Get baseline results"""
    baseline_models = create_baseline_models()
    baseline_results = pd.DataFrame(columns=metrics_columns)
    for index, model in enumerate(baseline_models):
        history = model.fit(x_train, y_train,
                            epochs=FLAGS.epochs,
                            validation_data=(x_test, y_test),
                            batch_size=FLAGS.batch_size)
        if index == 0:
            m = "DNN-NO-DP"
        else:
            m = "LR-NO-DP"
        new_row = {'model': m,
                   'optimizer': 'SGD-NO-DP',
                   'epsilon': "INFINITY",
                   'noise_multiplier': "0",
                   'l2_norm_clip': "0",
                   'acc': round(history.history['accuracy'][-1], 4),
                   'val_acc': round(history.history['val_accuracy'][-1], 4),
                   'loss': round(history.history['loss'][-1], 4),
                   'val_loss': round(history.history['val_loss'][-1], 4),
                   'batch_size': FLAGS.batch_size,
                   'microbatches': FLAGS.microbatches,
                   'learning_rate': FLAGS.learning_rate,
                   }

        log_metrics_baseline(history, index)
        baseline_results.loc[len(baseline_results)] = new_row
    if "Baseline" not in book.sheetnames:
        book.create_sheet("Baseline")
    baseline_results.to_excel(writer, sheet_name="Baseline", startrow=writer.sheets["Baseline"].max_row,
                              index=True, header=True)

    """ Grid Search """
    model_grid = []

    aVals = [.25, .5, 1, 1.25, 1.5, 2]

    for norm_clip in aVals:
        for noise_mul in aVals:
            model_grid.append({
                'l2_norm_clip': FLAGS.l2_norm_clip * norm_clip,
                'noise_multiplier': FLAGS.noise_multiplier * noise_mul,
                'epochs': FLAGS.epochs,
                'acc': 0,
                'val_acc': 0,
                'loss': 0,
                'val_loss': 0,
                'optimizer': "none"
            })

    for optimizer_index in range(0, 2):
        for i, m in enumerate(model_grid):
            m['optimizer'] = str(optimizer_index)
            calculate_model(
                x_train, x_test, y_train, y_test, m, i
            )

    # calculate epsilon
    if env.bool("DP", False) is True:
        privacy_results = pd.DataFrame(columns=metrics_columns)
        for m in model_grid:
            epsilon = compute_epsilon_noise(FLAGS.epochs * 180000 // FLAGS.batch_size, m['noise_multiplier'])
            new_row = {'model': "DNN-DP",
                       'optimizer': m['optimizer'],
                       'epsilon': epsilon,
                       'noise_multiplier': m['noise_multiplier'],
                       'l2_norm_clip': m['l2_norm_clip'],
                       'acc': m['acc'],
                       'val_acc': m['val_acc'],
                       'loss': m['loss'],
                       'val_loss': m['val_loss'],
                       'batch_size': FLAGS.batch_size,
                       'microbatches': FLAGS.microbatches,
                       'learning_rate': FLAGS.learning_rate,
                       }

            privacy_results.loc[len(privacy_results)] = new_row
            print("Computed epsilon with l2_norm_clip = " + str(m['l2_norm_clip']) + ", noise_multiplier = " +
                  str(m['noise_multiplier']) + " is epsilon = " + str(epsilon))

        models = privacy_results['model'].unique()
        optimizers = privacy_results['optimizer'].unique()
        for m in models:
            for o in optimizers:
                if m not in book.sheetnames:
                    book.create_sheet(m + "_" + o)
                actual = privacy_results[(privacy_results["model"] == m) & (privacy_results["optimizer"] == o)]
                actual.to_excel(writer, sheet_name=m, startrow=writer.sheets[m].max_row, index=True, header=True)

        writer.save()

        l2_norm_clip_values = privacy_results['l2_norm_clip'].unique()
        print(l2_norm_clip_values)
        for u in l2_norm_clip_values:
            actual = privacy_results[privacy_results["l2_norm_clip"] == u]
            # sorted_actual = sorted(actual, key=lambda x: x["val_acc"], reverse=True)
            print(actual)
            # plot epsilon graph
            plt.plot(actual["noise_multiplier"], actual["epsilon"])
            plt.title('model epsilon values for l2_norm_clip ' + str(u))
            plt.ylabel('epsilon')
            plt.xlabel('noise_multiplier')
            plt.grid(True)
            plt.tight_layout()
            plt.savefig(actual_dir + "epsilon_l2_norm_clip_" + str(u).replace(".", '_'), bbox_inches='tight')
            logger.log({"Epsilon plot l2 = " + str(u): plt})
            plt.show()
            # plot accuracy graph
            plt.plot(actual["noise_multiplier"], actual["val_acc"])
            plt.plot(actual["noise_multiplier"], actual["val_loss"])
            plt.title('model val_acc, val_loss values for l2_norm_clip ' + str(u))
            plt.ylabel('val_acc, val_loss')
            plt.xlabel('noise_multiplier')
            plt.grid(True)
            plt.tight_layout()
            plt.savefig(actual_dir + "noise_accuracy_l2_norm_clip_" + str(u).replace(".", '_'), bbox_inches='tight')
            logger.log({"Accuracy by noise and clip plot l2 = " + str(u): plt})
            plt.show()

    wandb.finish()


if __name__ == '__main__':
    main()
